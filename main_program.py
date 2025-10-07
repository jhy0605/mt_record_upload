import os.path
import os.path
import re
import shutil
import zipfile
import openpyxl
import paramiko
import win32cred
import time
import datetime
from dingding_remind import DingTalkOAForwarder, Dingding_Warning
from db import open_sql


# 更新windows凭据
def add_windows_credential():
    voucher_disk = {'10.10.100.203': ['record_syn', 'Admin@1234'],
                    '10.10.100.204': ['jw\jianghongyu', 'Jw.dh@@@@2222'],
                    '10.10.100.83': ['record_syn', 'Admin@1234'],
                    '10.10.100.211': ['okccsmb', 'Admin@1234'],
                    '172.20.50.249': ['dhvoip', 'Admin@1234'],
                    '192.168.106.249': ['mfvoip', 'Admin@1234'],
                    '10.10.100.209': ['record_syn', 'Admin@1234'],
                    }

    for share in voucher_disk.keys():
        u_list = voucher_disk[share]
        username = u_list[0]
        password = u_list[1]
        # 准备凭证信息
        credential = {
            'TargetName': share,
            'Type': win32cred.CRED_TYPE_DOMAIN_PASSWORD,
            'UserName': username,
            'CredentialBlob': password,
            'Persist': win32cred.CRED_PERSIST_LOCAL_MACHINE  # 或者使用 win32cred.CRED_PERSIST_ENTERPRISE，根据需要选择
        }
        # 添加凭证
        try:
            win32cred.CredWrite(credential, 0)
            print('{}凭据添加成功'.format(share))
        except Exception as k:
            print('{}凭据添加失败'.format(share))


# 录音匹配案件，更新数据库
class real_time_upload:
    def __init__(self):
        # 案件表参数
        self.data_path = ''
        self.cas_file_path_a = r'\\10.10.100.204\文件共享\互金\美团项目\技术部共享\原始数据'
        # self.cas_file_path_b = r'D:\BaiduSyncdisk\python\develop\mt_record_upload\table'

    # 读取案件表，输出案件与号码列表
    def out_bank_list(self):

        print('开始读取案件表')
        list_phone_all = []
        bank_error = ''
        list_a = os.listdir(self.cas_file_path_a)
        for a in list_a:
            file_a = os.path.join(self.cas_file_path_a, a)  # 名字目录
            print(a)
            if os.path.isdir(file_a):
                list_b = os.listdir(file_a)  # 表格
                for b in list_b:
                    if os.path.splitext(b)[1] == ".xlsx" and '~$' not in b:
                        cas_file_path = os.path.join(file_a, b)
                        if os.path.exists(cas_file_path):
                            print(cas_file_path)
                            wd = openpyxl.load_workbook(cas_file_path)
                            ws = wd.active
                            max_row = ws.max_row
                            cells = ws['A2:I{}'.format(max_row)]
                            for cell in cells:
                                phone_list = []  # 电话号码列表
                                if cell[0].value:  # 如果案件号存在
                                    # 添加手机号、联系人手机号
                                    for phone_a in [cell[2].value, cell[4].value]:
                                        if phone_a:
                                            phone_list.append(str(phone_a).replace('-', ''))
                                    # 添加联系信息1、联系信息2
                                    for phone in [cell[5].value, cell[6].value]:
                                        if phone:
                                            phone = str(phone)
                                            if len(phone) > 5:
                                                p = re.compile('_\d+_')
                                                phone_s_list = p.findall(phone)
                                                if len(phone_s_list) > 0:
                                                    for phone_s in phone_s_list:
                                                        phone_list.append(phone_s.replace('_', ''))
                                    # print(cell[0].value, phone_list)
                                    xm = cell[8].value
                                    if not xm or '=' in xm:
                                        bank_error += (
                                            '案件表：{} 从第{}行开始有格式错误，请及时修正！\n'.format(b, cell[8].row))
                                        break
                                    list_value = [str(cell[0].value).replace(' ', '').replace('\n', ''), phone_list,
                                                  xm.replace('\n', '')]
                                    list_phone_all.append(list_value)
                            wd.close()
        if bank_error:
            bank_error += '请注意：案件表中禁止包含公式，包含公式的行将无法识别！'
            print(bank_error)
            DingTalkOAForwarder().ding_sampletext(bank_error)  # 发送文字

        return list_phone_all

    # 录音匹配案件，更新数据库
    def mat_bank(self, date):

        # 查询录音
        record_list = open_sql().select_data(date)
        # 查询案件
        bank_list = real_time_upload().out_bank_list()

        # 判断如果有录音和案件，进行匹配
        if len(record_list) > 0 and len(bank_list) > 0:
            for record in record_list:
                # 单条录音
                for bank in bank_list:
                    if record[4] in bank[1]:  # 匹配成功
                        # 更新数据库
                        open_sql().update_table('case_id', bank[0], record[0])  # 更新案件号
                        open_sql().update_table('product', bank[2], record[0])  # 更新产品名称
                        break


# 压缩、加密、上传录音
class up_records:
    def __init__(self):
        self.data_path = ''  # 文件存放位置
        self.str_time = ''  # 时间字符串

        # sftp参数
        self.sftp_ip = 'one-sftp.sankuai.com'
        self.sftp_port = 2222
        self.sftp_user = 'collection-jinw'
        self.sftp_passwd = 'hgGZS3QzqIDY5l'
        self.bz_sftp_path = r'/onesftp-collection/data/audio/jinw/biaozhun'
        self.fbz_sftp_path = r'/onesftp-collection/data/audio/jinw/feibiaozhun'
        # 催记上传目录：/onesftp-collection/data/record/jinw

    # copy录音，生成映射表
    def create_excle(self, record_list):
        # 创建文件夹
        record_path = os.path.join(self.data_path, 'data', self.str_time)  # 录音存放目录
        if os.path.exists(record_path):
            shutil.rmtree(record_path)
        os.makedirs(record_path)

        # 新建一个表格，为索引文件
        wd = openpyxl.Workbook()
        ws = wd.active
        ws.append(['文件名', '外呼时间', '被叫号码', '客户号', '产品名称', '主叫号码'])
        if len(record_list) > 0:  # 判断如果有录音和案件
            for record in record_list:
                # 索引文件
                index = [record[1], record[3].strftime('%Y-%m-%d %H:%M:%S'), record[4], record[6], record[7], 'unkown']
                print(index)
                ws.append(index)
                # copy录音
                shutil.copy(os.path.join(record[2], record[1]), os.path.join(record_path, record[1]))

        wd.save(os.path.join(record_path, 'yingshebiao.xlsx'))

    # 压缩文件为.zip格式, 加密文件
    def ccmsApp(self):
        # 压缩文件
        record_path = os.path.join(self.data_path, 'data', self.str_time)
        os.chdir(record_path)
        zip_path = os.path.join(record_path, self.str_time + '.zip')  # 压缩文件
        z = zipfile.ZipFile(zip_path, 'w')
        for record in os.listdir(record_path):
            if os.path.splitext(record)[1] != '.zip':
                z.write(record)
                os.remove(record)
        z.close()
        print('压缩文件成功')

        # 加密文件
        print('开始加密文件')
        tool_path = os.path.join(self.data_path, 'tool')  # 加解密工具存放位置
        cmd = r'{}\\bin\\java -jar {}\\tool.jar 加密 {}'.format(tool_path, tool_path, zip_path)
        print(cmd)
        os.system(cmd)

    # 上传录音
    def up_records(self, d_path):
        record_path = os.path.join(self.data_path, 'data', self.str_time)
        os.chdir(os.path.dirname(record_path))  # 切换工作目录，取消文件占用
        print('开始连接sftp服务器')
        t = paramiko.Transport((self.sftp_ip, self.sftp_port))
        t.connect(username=self.sftp_user, password=self.sftp_passwd)
        sftp = paramiko.SFTPClient.from_transport(t)

        file_list = sftp.listdir(d_path)
        date = datetime.datetime.now().strftime('%Y%m%d')
        if date not in file_list:
            sftp.mkdir(d_path + '/' + date)
        print('上传文件：', os.path.join(record_path, self.str_time + '_加密.zip'))
        sftp.put(os.path.join(record_path, self.str_time + '_加密.zip'),
                 d_path + '/' + date + '/' + self.str_time + '.zip')
        print('文件上传成功')
        sftp.close()
        t.close()
        shutil.rmtree(record_path)
        time.sleep(1)

    # 非标准上传
    def f_run(self):
        # 计算昨天的日期
        yesterday = datetime.datetime.now() - datetime.timedelta(days=1)
        # 计算结束时间
        end_date = yesterday.date().strftime('%Y-%m-%d ') + '00:00:00'

        record_list = open_sql().select_data(end_date)  # 截止昨天未匹配案件的录音
        up_records.create_excle(self, record_list)  # copy录音，生成映射表
        up_records.ccmsApp(self)  # 压缩文件为.zip格式, 加密文件
        up_records.up_records(self, self.fbz_sftp_path)  # 非标准上传录音

        # 更新状态
        date_t = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for record in record_list:
            r_id = record[0]
            open_sql().update_table('upload_time', date_t, r_id)
            open_sql().update_table('mark', '非标准上传', r_id)

    #  标准上传录音
    def run(self):
        record_list = open_sql().select_up_data()  # 查询需要上传的录音数据
        up_records.create_excle(self, record_list)  # copy录音，生成映射表
        up_records.ccmsApp(self)  # 压缩文件为.zip格式, 加密文件
        up_records.up_records(self, self.bz_sftp_path)  # 标准上传录音

        # 更新数据库状态
        date_t = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for record in record_list:
            r_id = record[0]
            open_sql().update_table('upload_time', date_t, r_id)


# 录音上传监控
class status_check:
    def __init__(self):
        self.pwd_path = r'E:\mt_record_upload'
        # self.pwd_path = r'D:\BaiduSyncdisk\python\develop\mt_record_upload'

    #  查询未匹配案件的录音列表,生成未匹配案件录音汇总.xlsx
    def search_fail_record(self):
        file_path = './未匹配案件录音汇总.xlsx'
        end_tate = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        # 查询截止现在未匹配成功的录音
        select_record_list = open_sql().select_data(end_tate)
        n = len(select_record_list)
        wd = openpyxl.Workbook()
        ws = wd.active
        ws.append(['ID', '文件名', '路径', '时间', '联系号码', '呼叫系统', '案件号', '产品名称', '上传时间', '标注'])
        if n > 0:  # 判断如果有录音和案件
            for record in select_record_list:
                # 索引文件
                ws.append(record)
        wd.save(file_path)
        return file_path, n

    # 9点播报
    def remind_9(self):
        # 查询截止现在未匹配成功的录音，生成.xlsx和数量
        file_path, n = status_check.search_fail_record(self)

        # 查询今天的录音上传情况
        sj = open_sql().select_success_rate(1)
        if sj[0] == 0:
            sj.append(0)
            pl = '0%'
        else:
            pl = "{:.0%}".format(sj[1] / sj[0])
        text = "昨日美团录音上传数据统计：\n" \
               "产生录音数量：{}\n" \
               "上传录音数量：{}\n" \
               "录音案件匹配率：{}\n" \
               "----------------------------\n" \
               "截止现在未匹配案件录音数量：{}\n" \
               "请及时更新案件表!".format(sj[0], sj[1], pl, n)  # ，昨天未匹配案件的录音今天晚上21点将进行失效标记

        DingTalkOAForwarder().ding_sampletext(text)  # 发送文字
        mediaid, filename = DingTalkOAForwarder().get_file(file_path)  # 上传文件，得到mediaid和文件名
        DingTalkOAForwarder().ding_samplefile(mediaid, filename)  # 发送文件

        if os.path.exists(file_path):
            os.remove(file_path)

    # 16点播报
    def remind_16(self):
        # 查询截止现在未匹配成功的录音，生成.xlsx和数量
        file_path, n = status_check.search_fail_record(self)
        # 查询今天的录音上传情况
        sj = open_sql().select_success_rate(0)
        if sj[0] == 0:
            sj.append(0)
            pl = '0%'
        else:
            pl = "{:.0%}".format(sj[1] / sj[0])
        text = "今日（截止现在）美团录音上传数据统计：\n" \
               "产生录音数量：{}\n" \
               "上传录音数量：{}\n" \
               "录音案件匹配率：{}\n" \
               "----------------------------\n" \
               "截止现在未匹配案件录音：{}\n" \
               "请及时更新案件表!\n" \
               "未匹配案件录音将在今天晚上21点进行非标准上传。" \
            .format(sj[0], sj[1], pl, n)

        DingTalkOAForwarder().ding_sampletext(text)  # 发送文字
        mediaid, filename = DingTalkOAForwarder().get_file(file_path)  # 上传文件，得到mediaid和文件名
        DingTalkOAForwarder().ding_samplefile(mediaid, filename)  # 发送文件

        if os.path.exists(file_path):
            os.remove(file_path)

    # 21点 对截止前一天未上传的所有录音进行非标准上传。
    def remind_21(self):
        UR = up_records()
        UR.data_path = self.pwd_path
        UR.str_time = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        UR.f_run()  # 非标准上传

    # 定时钉钉播报程序
    def run(self):
        add_windows_credential()
        t = n = 0
        while t == 0:
            if n < 300:  # 五分钟没有运行就自动关闭
                time_1 = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                h = time_1[8:10]
                m = time_1[10:12]
                s = time_1[12:14]
                print(time_1)
                if h == '09' and m == '00' and s == '00':
                    status_check().remind_9()
                    t += 1
                elif h == '16' and m == '00' and s == '00':
                    status_check().remind_16()
                    t += 1
                elif h == '21' and m == '00' and s == '00':
                    status_check().remind_21()
                    t += 1
                else:
                    time.sleep(1)
                    n += 1
            else:
                return None
